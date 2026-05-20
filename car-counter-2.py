from ultralytics import YOLO
import cv2
import cvzone
import math
import numpy as np
import time
from sort import *
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE_PATH = "assets/"
IMAGES_PATH = BASE_PATH + "images/"
VIDEOS_PATH = BASE_PATH + "videos/"
WEIGHTS_PATH = BASE_PATH + "Yolo-Weights/"
MASK_PATH = BASE_PATH + "mask/"

# ─── SPEED CALIBRATION ──────────────────────────────────────────────────────
# Tune these two values for your specific camera/road setup:
#   REAL_WORLD_DISTANCE_M  = the actual distance (in metres) that one pixel
#                            represents at the counting-line depth.
#   FPS                    = frame-rate of the source video.
#
# Quick calibration: if you know a car is ~4.5 m long and its bounding box is
# ~120 px wide at the line, then 1 px ≈ 4.5/120 ≈ 0.0375 m.
PIXEL_TO_METER = 0.05          # metres per pixel (adjust for your camera)
VIDEO_FPS      = 30            # source video FPS
SPEED_SMOOTH   = 5             # frames to average for smoother speed

# ─── VIDEO SETUP ────────────────────────────────────────────────────────────
cap = cv2.VideoCapture(VIDEOS_PATH + "car2.mp4")
cap.set(3, 1280)
cap.set(4, 720)

# ─── YOLO CLASSES ───────────────────────────────────────────────────────────
yolo_classes = {
    0: "person", 1: "bicycle", 2: "car", 3: "motorcycle", 4: "airplane",
    5: "bus", 6: "train", 7: "truck", 8: "boat", 9: "traffic light",
    10: "fire hydrant", 11: "stop sign", 12: "parking meter", 13: "bench",
    14: "bird", 15: "cat", 16: "dog", 17: "horse", 18: "sheep", 19: "cow",
    20: "elephant", 21: "bear", 22: "zebra", 23: "giraffe", 24: "backpack",
    25: "umbrella", 26: "handbag", 27: "tie", 28: "suitcase", 29: "frisbee",
    30: "skis", 31: "snowboard", 32: "sports ball", 33: "kite",
    34: "baseball bat", 35: "baseball glove", 36: "skateboard",
    37: "surfboard", 38: "tennis racket", 39: "bottle", 40: "wine glass",
    41: "cup", 42: "fork", 43: "knife", 44: "spoon", 45: "bowl",
    46: "banana", 47: "apple", 48: "sandwich", 49: "orange", 50: "broccoli",
    51: "carrot", 52: "hot dog", 53: "pizza", 54: "donut", 55: "cake",
    56: "chair", 57: "couch", 58: "potted plant", 59: "bed",
    60: "dining table", 61: "toilet", 62: "tv", 63: "laptop", 64: "mouse",
    65: "remote", 66: "keyboard", 67: "cell phone", 68: "microwave",
    69: "oven", 70: "toaster", 71: "sink", 72: "refrigerator", 73: "book",
    74: "clock", 75: "vase", 76: "scissors", 77: "teddy bear",
    78: "hair drier", 79: "toothbrush",
}

VEHICLE_CLASSES = {"car", "truck", "bus", "motorcycle"}

# ─── MODEL & TRACKER ────────────────────────────────────────────────────────
model   = YOLO(WEIGHTS_PATH + "yolov8n.pt")
mask    = cv2.imread(MASK_PATH + "car-mask.png")
tracker = Sort(max_age=20, min_hits=3, iou_threshold=0.3)

# Counting line [x1, y1, x2, y2]
limits = [100, 350, 850, 200]

# ─── TRACKING STATE ─────────────────────────────────────────────────────────
total_count   = []          # list of IDs that crossed the line
prev_centers  = {}          # {id: deque of (frame_no, cx, cy)}
speed_history = {}          # {id: deque of speed samples}
car_speeds    = {}          # {id: final speed km/h when it crossed the line}
car_classes   = {}          # {id: vehicle class label}
car_timestamps= {}          # {id: wall-clock timestamp string}
frame_no      = 0

# ─── HELPERS ────────────────────────────────────────────────────────────────
def estimate_speed(track_id):
    """Return smoothed speed in km/h from recent centre positions."""
    if track_id not in prev_centers or len(prev_centers[track_id]) < 2:
        return 0.0
    pts = list(prev_centers[track_id])
    f0, x0, y0 = pts[0]
    f1, x1, y1 = pts[-1]
    dt_frames = f1 - f0
    if dt_frames == 0:
        return 0.0
    dist_px   = math.hypot(x1 - x0, y1 - y0)
    dist_m    = dist_px * PIXEL_TO_METER
    dt_sec    = dt_frames / VIDEO_FPS
    speed_ms  = dist_m / dt_sec
    return speed_ms * 3.6   # → km/h


def save_excel(car_speeds, car_classes, car_timestamps, out_path="car_speeds.xlsx"):
    """Write results to a nicely formatted Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Speeds"

    # ── colour palette ──
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    alt_fill  = PatternFill("solid", start_color="D6E4F0")
    thin      = Side(style="thin", color="AAAAAA")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Arial", size=10)
    ctr       = Alignment(horizontal="center", vertical="center")

    # ── headers ──
    headers = ["#", "Vehicle ID", "Class", "Speed (km/h)", "Timestamp"]
    col_widths = [6, 14, 16, 18, 26]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = ctr
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    # ── data rows ──
    sorted_ids = sorted(car_speeds.keys())
    for row_idx, vid in enumerate(sorted_ids, start=2):
        row_data = [
            row_idx - 1,
            vid,
            car_classes.get(vid, "unknown"),
            round(car_speeds[vid], 1),
            car_timestamps.get(vid, ""),
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = body_font
            cell.alignment = ctr
            cell.border    = border
            if fill:
                cell.fill = fill

    # ── summary row ──
    summary_row = len(sorted_ids) + 2
    ws.cell(row=summary_row, column=1, value="Total").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=summary_row, column=2, value=f'=COUNTA(B2:B{summary_row-1})').font = Font(name="Arial", bold=True)
    ws.cell(row=summary_row, column=4, value=f'=AVERAGE(D2:D{summary_row-1})').font = Font(name="Arial", bold=True)
    ws.cell(row=summary_row, column=4).number_format = "0.0"
    for col in range(1, 6):
        ws.cell(row=summary_row, column=col).border = border
        ws.cell(row=summary_row, column=col).alignment = ctr

    # ── stats sheet ──
    ws2 = wb.create_sheet("Summary Stats")
    stats = [
        ("Metric", "Value"),
        ("Total Vehicles", f"=COUNTA('Vehicle Speeds'!B2:B{summary_row-1})"),
        ("Avg Speed (km/h)", f"=AVERAGE('Vehicle Speeds'!D2:D{summary_row-1})"),
        ("Max Speed (km/h)", f"=MAX('Vehicle Speeds'!D2:D{summary_row-1})"),
        ("Min Speed (km/h)", f"=MIN('Vehicle Speeds'!D2:D{summary_row-1})"),
        ("Cars",        f"=COUNTIF('Vehicle Speeds'!C2:C{summary_row-1},\"car\")"),
        ("Trucks",      f"=COUNTIF('Vehicle Speeds'!C2:C{summary_row-1},\"truck\")"),
        ("Buses",       f"=COUNTIF('Vehicle Speeds'!C2:C{summary_row-1},\"bus\")"),
        ("Motorcycles", f"=COUNTIF('Vehicle Speeds'!C2:C{summary_row-1},\"motorcycle\")"),
    ]
    for r, (label, val) in enumerate(stats, 1):
        for c, v in enumerate([label, val], 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.font   = Font(name="Arial", bold=(r == 1), size=10)
            cell.border = border
            cell.alignment = ctr
        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 18

    wb.save(out_path)
    print(f"[✓] Excel saved → {out_path}")


# ─── MAIN LOOP ──────────────────────────────────────────────────────────────
while True:
    success, img = cap.read()
    if not success:
        break
    frame_no += 1

    imgRegion  = cv2.bitwise_and(img, mask)
    imgGraphics = cv2.imread(IMAGES_PATH + "car-counter-graphic.png", cv2.IMREAD_UNCHANGED)
    h, w, _    = imgGraphics.shape
    img[0:h, 0:w] = imgGraphics

    results    = model(imgRegion, stream=True)
    detections = np.empty((0, 5))
    det_classes= {}   # temp map bbox → class (approximation)

    for r in results:
        for box in r.boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            conf = math.ceil(box.conf[0] * 100) / 100
            cls  = yolo_classes[int(box.cls[0])]
            if cls in VEHICLE_CLASSES and conf > 0.2:
                detections = np.vstack((detections, [x1, y1, x2, y2, conf]))
                det_classes[(x1, y1, x2, y2)] = cls

    # Draw counting line
    cv2.line(img, (limits[0], limits[1]), (limits[2], limits[3]), (0, 0, 255), 5)

    results_tracker = tracker.update(detections)

    for result in results_tracker:
        x1, y1, x2, y2, _id = map(int, result)
        cx, cy = (x1 + x2) // 2, (y1 + y2) // 2

        # Update position history for speed
        if _id not in prev_centers:
            from collections import deque
            prev_centers[_id]  = deque(maxlen=SPEED_SMOOTH)
            speed_history[_id] = deque(maxlen=10)
        prev_centers[_id].append((frame_no, cx, cy))

        spd = estimate_speed(_id)
        speed_history[_id].append(spd)
        avg_spd = sum(speed_history[_id]) / len(speed_history[_id])

        # Best-guess class: find closest detection
        best_cls = "vehicle"
        best_dist = float("inf")
        for (dx1, dy1, dx2, dy2), cls_name in det_classes.items():
            dist = abs(cx - (dx1 + dx2) // 2) + abs(cy - (dy1 + dy2) // 2)
            if dist < best_dist:
                best_dist, best_cls = dist, cls_name
        if _id not in car_classes:
            car_classes[_id] = best_cls

        # Draw bounding box + speed label
        cv2.rectangle(img, (x1, y1), (x2, y2), (255, 0, 0), 2)
        cvzone.putTextRect(
            img, f"ID:{_id} {avg_spd:.0f}km/h",
            (max(x1, 0), max(y1 - 10, 35)),
            scale=0.6, thickness=1, offset=2,
        )
        cv2.circle(img, (cx, cy), 4, (0, 255, 255), cv2.FILLED)

        # Check line crossing
        if limits[0] < cx < limits[2] and limits[1] - 15 < cy < limits[1] + 15:
            cv2.line(img, (limits[0], limits[1]), (limits[2], limits[3]), (0, 255, 0), 5)
            if _id not in total_count:
                total_count.append(_id)
                car_speeds[_id]     = avg_spd
                car_timestamps[_id] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                car_classes[_id]    = best_cls

    # HUD
    cvzone.putTextRect(
        img, f"Count: {len(total_count)}", (50, 33),
        scale=2, thickness=4, offset=3, colorR=(205, 199, 149),
    )

    cv2.imshow("AutoCount AI – Speed Tracker", img)
    if cv2.waitKey(1) == 27:   # ESC to quit early
        break

cap.release()
cv2.destroyAllWindows()

# ─── EXPORT TO EXCEL ────────────────────────────────────────────────────────
if car_speeds:
    save_excel(car_speeds, car_classes, car_timestamps, "car_speeds.xlsx")
else:
    print("[!] No vehicles crossed the line – Excel not saved.")